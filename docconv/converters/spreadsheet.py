# docconv/converters/spreadsheet.py
from __future__ import annotations

import csv
import re
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any

import chardet
import openpyxl
import pandas as pd

from docconv.config import Config, SpreadsheetConfig
from docconv.converters.base import BaseConverter

warnings.filterwarnings("ignore", category=UserWarning)


class SpreadsheetConverter(BaseConverter):
    EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}

    def can_handle(self, path: Path) -> bool:
        return path.suffix.lower() in self.EXTENSIONS

    def convert(self, path: Path, config: Config) -> str:
        suffix = path.suffix.lower()
        sc = config.spreadsheet
        if suffix in (".xlsx", ".xls", ".xlsm"):
            return self._convert_excel(path, sc)
        return self._convert_csv(path, sc)

    # ── Excel ────────────────────────────────────────────────────

    def _convert_excel(self, path: Path, sc: SpreadsheetConfig) -> str:
        sections: list[str] = []
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
            target = sc.sheets or wb.sheetnames
            for name in target:
                if name not in wb.sheetnames:
                    continue
                grid = self._fill_merged(wb[name])
                df = self._grid_to_df(grid, sc)
                sections.append(self._render_section(name, df))
        except Exception:
            xl = pd.ExcelFile(str(path))
            target = sc.sheets or xl.sheet_names
            for name in target:
                df_raw = xl.parse(name, header=None, dtype=str).fillna("")
                grid = df_raw.values.tolist()
                df = self._grid_to_df(grid, sc)
                sections.append(self._render_section(name, df))
        return "\n".join(s for s in sections if s)

    def _fill_merged(self, ws: Any) -> list[list[Any]]:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        grid: list[list[Any]] = [[None] * max_col for _ in range(max_row)]
        for row in ws.iter_rows():
            for cell in row:
                grid[cell.row - 1][cell.column - 1] = cell.value
        for rng in ws.merged_cells.ranges:
            top = grid[rng.min_row - 1][rng.min_col - 1]
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    grid[r][c] = top
        return grid

    def _render_section(self, name: str, df: pd.DataFrame) -> str:
        if df.empty:
            return ""
        table = df.to_markdown(index=False, tablefmt="pipe") or "_empty_"
        return f"## {name}\n\n{table}\n"

    # ── CSV ──────────────────────────────────────────────────────

    def _convert_csv(self, path: Path, sc: SpreadsheetConfig) -> str:
        encoding = self._detect_encoding(path)
        delimiter = self._detect_delimiter(path, encoding)
        df_raw = pd.read_csv(
            path, encoding=encoding, sep=delimiter,
            header=None, dtype=str, keep_default_na=False, engine="python",
        )
        df = self._grid_to_df(df_raw.values.tolist(), sc)
        if df.empty:
            return "_empty_\n"
        return df.to_markdown(index=False, tablefmt="pipe") + "\n"

    def _detect_encoding(self, path: Path) -> str:
        raw = path.read_bytes()
        if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
            return "utf-16"
        if raw.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        return chardet.detect(raw[:65_536]).get("encoding") or "utf-8"

    def _detect_delimiter(self, path: Path, encoding: str) -> str:
        sample = path.read_text(encoding=encoding, errors="replace")[:8192]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            counts = {d: sample.count(d) for d in ",;\t|"}
            return max(counts, key=lambda k: counts[k])

    # ── Shared helpers ───────────────────────────────────────────

    def _grid_to_df(self, grid: list[list[Any]], sc: SpreadsheetConfig) -> pd.DataFrame:
        if not grid:
            return pd.DataFrame()
        n_cols = max((len(r) for r in grid), default=0)
        grid = [row + [None] * (n_cols - len(row)) for row in grid]

        hr = sc.header_rows
        if hr <= 0:
            headers = [f"Col{i}" for i in range(n_cols)]
            body = grid
        elif hr == 1:
            headers = self._make_headers(grid[0])
            body = grid[1:]
        else:
            flat = []
            for ci in range(n_cols):
                parts = [
                    str(grid[ri][ci]).strip()
                    for ri in range(hr)
                    if grid[ri][ci] is not None and str(grid[ri][ci]).strip()
                ]
                flat.append(" › ".join(parts) if parts else f"Col{ci}")
            headers = self._make_headers(flat)
            body = grid[hr:]

        df = pd.DataFrame(body, columns=headers)
        try:
            df = df.map(lambda v: self._fmt(v, sc.max_cell_len))
        except AttributeError:
            df = df.applymap(lambda v: self._fmt(v, sc.max_cell_len))

        df = df.replace("", pd.NA)
        if sc.skip_empty_sheets:
            df = df.dropna(how="all").dropna(axis=1, how="all")
        return df.fillna("").reset_index(drop=True)

    def _make_headers(self, raw: list[Any]) -> list[str]:
        seen: dict[str, int] = {}
        result = []
        for i, h in enumerate(raw):
            key = str(h).strip() if h is not None and str(h).strip() else f"Col{i}"
            if key in seen:
                seen[key] += 1
                result.append(f"{key}_{seen[key]}")
            else:
                seen[key] = 0
                result.append(key)
        return result

    def _fmt(self, val: Any, max_len: int) -> str:
        if val is None:
            return ""
        if isinstance(val, bool):
            return "TRUE" if val else "FALSE"
        if isinstance(val, datetime):
            text = val.strftime("%Y-%m-%d %H:%M:%S") if (val.hour or val.minute or val.second) else val.strftime("%Y-%m-%d")
        elif isinstance(val, date):
            text = val.strftime("%Y-%m-%d")
        elif isinstance(val, float):
            if val != val:
                return ""
            text = str(int(val)) if val.is_integer() else f"{val:g}"
        else:
            text = str(val)
        text = text.strip()
        text = re.sub(r"[\r\n]+", " ", text)
        text = text.replace("|", "\\|")
        if max_len > 0 and len(text) > max_len:
            text = text[: max_len - 1] + "…"
        return text
